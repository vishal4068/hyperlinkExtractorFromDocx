from docx import Document
from openpyxl import Workbook
from lxml import etree
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
# Load the Word document
doc = Document('vishal.docx')

# Create a new Excel workbook
workbook = Workbook()
worksheet = workbook.active

# Parse the XML
xml_content = etree.fromstring(doc.part.blob)

# Find hyperlinks
hyperlinks = xml_content.xpath('//w:hyperlink', namespaces={'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'})

# Add headers
worksheet.append(['Hyperlink Text', 'Hyperlink URL'])

# Set header row formatting
for cell in worksheet[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
hyperlink_count=0
# Initialize variables to store maximum content widths
max_text_width = 0
max_url_width = 0
# Iterate through hyperlinks
for hyperlink in hyperlinks:
    hyperlink_text = ''
    for elem in hyperlink.iter():
        if elem.tag.endswith('t') or elem.tag.endswith('delText'):
            if elem.text:
                hyperlink_text += elem.text
            if elem.tail:
                hyperlink_text += elem.tail
        elif elem.tag.endswith('softHyphen') or elem.tag.endswith('noBreakHyphen'):
            hyperlink_text += '-'  # For hyphen characters
    hyperlink_rid = hyperlink.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
    hyperlink_url = doc.part.rels[hyperlink_rid].target_ref
    # Update maximum content widths
    max_text_width = max(max_text_width, len(hyperlink_text))
    max_url_width = max(max_url_width, len(hyperlink_url))

    worksheet.append([hyperlink_text, f'=HYPERLINK("{hyperlink_url}", "{hyperlink_url}")'])
    hyperlink_count += 1 

for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.font=Font(underline='single', color='0000FF')

# Set column widths based on maximum content widths
worksheet.column_dimensions['A'].width = max_text_width + 1  # Add some padding
worksheet.column_dimensions['B'].width = max_url_width + 1   # Add some padding
# Save the Excel file
print("Extracted!!! Please Check in Current Folder")
print("Hyperlinks Count=", hyperlink_count)
workbook.save('extractedHyperlink.xlsx')